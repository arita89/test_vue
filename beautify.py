import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule


def create_workbook_with_data(df: pd.DataFrame, sheet_name: str = "Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    return wb, ws


def add_excel_table(ws, df: pd.DataFrame, table_name: str = "DataTable"):
    max_row = ws.max_row
    max_col = ws.max_column
    end_col_letter = get_column_letter(max_col)
    table_range = f"A1:{end_col_letter}{max_row}"

    table = Table(displayName=table_name, ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)


def format_cells(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    center_align = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = center_align
            if cell.row == 1:
                cell.font = bold_font


def autosize_columns(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2


def add_conditional_formatting(ws, df: pd.DataFrame, status_col: str, reference_col: str):
    if status_col not in df.columns or reference_col not in df.columns:
        return

    col_letters = {col: get_column_letter(idx + 1) for idx, col in enumerate(df.columns)}
    status_letter = col_letters[status_col]
    score_letter = col_letters[reference_col]
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        status_cell = f"{status_letter}{row}"
        score_cell = f"{score_letter}{row}"

        ws.conditional_formatting.add(status_cell,
            FormulaRule(
                formula=[f"${score_cell}>=80"],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            )
        )

        ws.conditional_formatting.add(status_cell,
            FormulaRule(
                formula=[f'AND(${score_cell}>=50, ${score_cell}<80)'],
                fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            )
        )

        ws.conditional_formatting.add(status_cell,
            FormulaRule(
                formula=[f"${score_cell}<50"],
                fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
            )
        )


def apply_hyperlinks(ws, df, link_col: str, label_prefix: str = "data of", label_col: str = None):
    if link_col not in df.columns:
        return

    link_col_idx = list(df.columns).index(link_col) + 1
    link_letter = get_column_letter(link_col_idx)
    label_col_idx = list(df.columns).index(label_col) + 1 if label_col and label_col in df.columns else None

    hyperlink_font = Font(color="0000FF", underline="single")

    for row_num in range(2, ws.max_row + 1):
        cell = ws[f"{link_letter}{row_num}"]
        url = cell.value
        if url:
            if label_col_idx:
                label_value = ws[f"{get_column_letter(label_col_idx)}{row_num}"].value
                label = f"{label_prefix} {label_value}"
            else:
                label = f"{label_prefix} {row_num - 1}"

            cell.value = label
            cell.hyperlink = url
            cell.font = hyperlink_font


def save_styled_excel(
    df: pd.DataFrame,
    filepath: str,
    table_name: str = "DataTable",
    sheet_name: str = "Sheet1",
    show_gridlines: bool = True,
    freeze_header: bool = True,
    format_status_column: bool = False,
    status_col: str = "Status",
    reference_col: str = "Score",
    add_links: bool = False,
    link_col: str = "Link",
    link_label_prefix: str = "data of",
    link_label_col: str = None
):
    wb, ws = create_workbook_with_data(df, sheet_name)
    add_excel_table(ws, df, table_name)
    format_cells(ws)
    autosize_columns(ws)

    if freeze_header:
        ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = show_gridlines

    if format_status_column:
        add_conditional_formatting(ws, df, status_col, reference_col)

    if add_links:
        apply_hyperlinks(ws, df, link_col, link_label_prefix, link_label_col)

    wb.save(filepath)

